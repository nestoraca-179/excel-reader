import os
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from models.AdmDTO import AdmDto
from models.ConDTO import ConDto


def validate_file(ruta: str) -> bool:
    """Valida si el archivo Excel existe"""
    if not os.path.exists(ruta):
        print(f"❌ ERROR: Archivo no encontrado: {ruta}")
        return False
    print(f"✅ Archivo encontrado: {ruta}")
    return True

def read_adm_sheet(ruta_excel: str) -> List[AdmDto]:
    """Lee hoja ADM y retorna lista de DTOs"""
    df_adm = None
    try:
        df_adm = pd.read_excel(ruta_excel, sheet_name='ADM', header=0)
        print(f"📊 Hoja ADM cargada: {len(df_adm)} filas")

        adm_list: List[AdmDto] = []
        for _, row in df_adm.iterrows():
            dto = AdmDto(
                nro_doc=str(row['nro_doc']) if pd.notna(row['nro_doc']) else '',
                co_tipo_doc=str(row['co_tipo_doc']) if pd.notna(row['co_tipo_doc']) else '',
                co_ven=str(row['co_ven']) if pd.notna(row['co_ven']) else '',
                co_cli=str(row['co_cli']) if pd.notna(row['co_cli']) else '',
                fec_emis=str(row['fec_emis']) if pd.notna(row['fec_emis']) else '',
                fec_venc=str(row['fec_venc']) if pd.notna(row['fec_venc']) else '',
                anulado=str(row['anulado']) if pd.notna(row['anulado']) else '',
                tasa=float(row['tasa']) if pd.notna(row['tasa']) else 0.0,
                total_neto=float(row['total_neto']) if pd.notna(row['total_neto']) else 0.0,
                saldo=float(row['saldo']) if pd.notna(row['saldo']) else 0.0,
                co_mone_doc=str(row['co_mone_doc']) if pd.notna(row['co_mone_doc']) else '',
                tasa_doc=float(row['tasa_doc']) if pd.notna(row['tasa_doc']) else 0.0,
                Rel_Inv=str(row['Rel_Inv']) if pd.notna(row['Rel_Inv']) else '',
                cli_des=str(row['cli_des']) if pd.notna(row['cli_des']) else '',
                observa=str(row['observa']) if pd.notna(row['observa']) else '',
                tipo_mov=str(row['tipo_mov']) if pd.notna(row['tipo_mov']) else '',
                Mon_Rep=str(row['Mon_Rep']) if pd.notna(row['Mon_Rep']) else '',
                Mon_Fil=float(row['Mon_Fil']) if pd.notna(row['Mon_Fil']) else 0.0,
                total_neto_new=float(row['total_neto_new']) if pd.notna(row['total_neto_new']) else 0.0,
                saldo_new=float(row['saldo_new']) if pd.notna(row['saldo_new']) else 0.0
            )
            adm_list.append(dto)

        return adm_list

    except FileNotFoundError:
        print("❌ ERROR: Hoja ADM no encontrada")
        return []
    except Exception as e:
        print(f"❌ ERROR leyendo ADM: {str(e)}")
        return []

def write_adm_sheet(ruta_excel: str, adm_list: List[AdmDto]):
    """
    Exporta lista de AdmDTO a Excel con formato profesional
    Hoja: 'RESULTADO_ADM'
    """
    try:
        print(f"\n📤 Exportando {len(adm_list):,} AdmDTOs a Excel...")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "RESULTADO_ADM"

        # 1. CABECERAS
        cabeceras = [
            'nro_doc', 'co_tipo_doc', 'co_ven', 'co_cli', 'fec_emis', 'fec_venc',
            'anulado', 'tasa', 'total_neto', 'saldo', 'co_mone_doc', 'tasa_doc',
            'Rel_Inv', 'cli_des', 'observa', 'tipo_mov', 'Mon_Rep', 'Mon_Fil',
            'total_neto_new', 'saldo_new', 'has_coincidence', 'row_coincidence', 'text_coincidence'
        ]

        # Escribir cabeceras
        for col, header in enumerate(cabeceras, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3673A5", end_color="3673A5", fill_type="solid")

        # 2. DATOS
        for fila, adm in enumerate(adm_list, 2):  # Fila 2 en adelante
            ws.cell(row=fila, column=1, value=adm.nro_doc)
            ws.cell(row=fila, column=2, value=adm.co_tipo_doc)
            ws.cell(row=fila, column=3, value=adm.co_ven)
            ws.cell(row=fila, column=4, value=adm.co_cli)
            ws.cell(row=fila, column=5, value=adm.fec_emis)
            ws.cell(row=fila, column=6, value=adm.fec_venc)
            ws.cell(row=fila, column=7, value=adm.anulado)
            ws.cell(row=fila, column=8, value=adm.tasa)
            ws.cell(row=fila, column=9, value=adm.total_neto)
            ws.cell(row=fila, column=10, value=adm.saldo)
            ws.cell(row=fila, column=11, value=adm.co_mone_doc)
            ws.cell(row=fila, column=12, value=adm.tasa_doc)
            ws.cell(row=fila, column=13, value=adm.Rel_Inv)
            ws.cell(row=fila, column=14, value=adm.cli_des)
            ws.cell(row=fila, column=15, value=adm.observa)
            ws.cell(row=fila, column=16, value=adm.tipo_mov)
            ws.cell(row=fila, column=17, value=adm.Mon_Rep)
            ws.cell(row=fila, column=18, value=adm.Mon_Fil)
            ws.cell(row=fila, column=19, value=adm.total_neto_new)
            ws.cell(row=fila, column=20, value=adm.saldo_new)
            ws.cell(row=fila, column=21, value=adm.has_coincidence)
            ws.cell(row=fila, column=22, value=adm.row_coincidence)
            ws.cell(row=fila, column=23, value=adm.text_coincidence)

        # 3. FORMATO PROFESIONAL
        # Ajustar ancho columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=23):
            for cell in row:
                cell.border = thin_border

        # 4. GUARDAR
        wb.save(ruta_excel)
        print(f"✅ EXCEL CREADO: '{ruta_excel}'")
        print(f"   📊 Filas: {len(adm_list):,}")
        print(f"   📋 Columnas: 23")
        print(f"   🎨 Hoja: RESULTADO_ADM")

        return ruta_excel

    except Exception as e:
        print(f"❌ Error exportando Excel: {str(e)}")
        return None

def read_adm_con(ruta_excel: str) -> List[ConDto]:
    """Lee hoja CON y retorna lista de DTOs"""
    df_con = None
    try:
        df_con = pd.read_excel(ruta_excel, sheet_name='CON', header=0)
        print(f"📊 Hoja CON cargada: {len(df_con)} filas")

        con_list: List[ConDto] = []
        for _, row in df_con.iterrows():
            dto = ConDto(
                co_cue=str(row['co_cue']) if pd.notna(row['co_cue']) else '',
                SaldoInicial=float(row['SaldoInicial']) if pd.notna(row['SaldoInicial']) else 0.0,
                MontoD=float(row['MontoD']) if pd.notna(row['MontoD']) else 0.0,
                MontoH=float(row['MontoH']) if pd.notna(row['MontoH']) else 0.0,
                EsActivo=str(row['EsActivo']) if pd.notna(row['EsActivo']) else '',
                EsPasivo=str(row['EsPasivo']) if pd.notna(row['EsPasivo']) else '',
                EsCapital=str(row['EsCapital']) if pd.notna(row['EsCapital']) else '',
                EsIngEgr=str(row['EsIngEgr']) if pd.notna(row['EsIngEgr']) else '',
                EsAdicional=str(row['EsAdicional']) if pd.notna(row['EsAdicional']) else '',
                detalle=str(row['detalle']) if pd.notna(row['detalle']) else '',
                des_cue=str(row['des_cue']) if pd.notna(row['des_cue']) else '',
                co_cuepadre=str(row['co_cuepadre']) if pd.notna(row['co_cuepadre']) else '',
                NivelCuenta=str(row['NivelCuenta']) if pd.notna(row['NivelCuenta']) else '',
                comp_num=str(row['comp_num']) if pd.notna(row['comp_num']) else '',
                fec_emis=str(row['fec_emis']) if pd.notna(row['fec_emis']) else '',
                descri=str(row['descri']) if pd.notna(row['descri']) else '',
                reng_num=str(row['reng_num']) if pd.notna(row['reng_num']) else '',
                docref=str(row['docref']) if pd.notna(row['docref']) else '',
                IncluirAsiento=str(row['IncluirAsiento']) if pd.notna(row['IncluirAsiento']) else '',
                SinCuentaMadre=str(row['SinCuentaMadre']) if pd.notna(row['SinCuentaMadre']) else '',
                debe_new=float(row['debe_new']) if pd.notna(row['debe_new']) else 0.0,
                haber_new=float(row['haber_new']) if pd.notna(row['haber_new']) else 0.0
            )
            con_list.append(dto)

        return con_list

    except FileNotFoundError:
        print("❌ ERROR: Hoja CON no encontrada")
        return []
    except Exception as e:
        print(f"❌ ERROR leyendo CON: {str(e)}")
        return []

def write_adm_con(ruta_excel: str, con_list: List[ConDto]):
    """
    Exporta lista de ConDTO a Excel con formato profesional
    Hoja: 'RESULTADO_CON'
    """
    try:
        print(f"\n📤 Exportando {len(con_list):,} ConDTOs a Excel...")

        wb = Workbook()
        ws = wb.active
        ws.title = "RESULTADO_CON"

        cabeceras = [
            'co_cue', 'SaldoInicial', 'MontoD', 'MontoH', 'EsActivo', 'EsPasivo',
            'EsCapital', 'EsIngEgr', 'EsAdicional', 'detalle', 'des_cue', 'co_cuepadre',
            'NivelCuenta', 'comp_num', 'fec_emis', 'descri', 'reng_num', 'docref',
            'IncluirAsiento', 'SinCuentaMadre', 'debe_new', 'haber_new'
        ]

        for col, header in enumerate(cabeceras, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3673A5", end_color="3673A5", fill_type="solid")

        for fila, con in enumerate(con_list, 2):
            ws.cell(row=fila, column=1, value=con.co_cue)
            ws.cell(row=fila, column=2, value=con.SaldoInicial)
            ws.cell(row=fila, column=3, value=con.MontoD)
            ws.cell(row=fila, column=4, value=con.MontoH)
            ws.cell(row=fila, column=5, value=con.EsActivo)
            ws.cell(row=fila, column=6, value=con.EsPasivo)
            ws.cell(row=fila, column=7, value=con.EsCapital)
            ws.cell(row=fila, column=8, value=con.EsIngEgr)
            ws.cell(row=fila, column=9, value=con.EsAdicional)
            ws.cell(row=fila, column=10, value=con.detalle)
            ws.cell(row=fila, column=11, value=con.des_cue)
            ws.cell(row=fila, column=12, value=con.co_cuepadre)
            ws.cell(row=fila, column=13, value=con.NivelCuenta)
            ws.cell(row=fila, column=14, value=con.comp_num)
            ws.cell(row=fila, column=15, value=con.fec_emis)
            ws.cell(row=fila, column=16, value=con.descri)
            ws.cell(row=fila, column=17, value=con.reng_num)
            ws.cell(row=fila, column=18, value=con.docref)
            ws.cell(row=fila, column=19, value=con.IncluirAsiento)
            ws.cell(row=fila, column=20, value=con.SinCuentaMadre)
            ws.cell(row=fila, column=21, value=con.debe_new)
            ws.cell(row=fila, column=22, value=con.haber_new)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        wb.save(ruta_excel)
        print(f"✅ EXCEL CREADO: '{ruta_excel}'")
        print(f"   📊 Filas: {len(con_list):,}")
        print(f"   📋 Columnas: {len(cabeceras)}")
        print(f"   🎨 Hoja: RESULTADO_CON")

        return ruta_excel

    except Exception as e:
        print(f"❌ Error exportando Excel (CON): {str(e)}")
        return None